import openpyxl

def getrowcount(file, SheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(SheetName)
    return (sheet.max_row)

def getColumnCount(file, SheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(SheetName)
    return (sheet.max_column)

def readData(file, SheetName, rownum, columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(SheetName)
    return (sheet.cell(row=rownum, column= columnnum).value)

def writedata(file, SheetName, rownum, columnnum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(SheetName)
    sheet.cell(row=rownum, column=columnnum).value = data
    workbook.save(file)





